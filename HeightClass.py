# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl
import math

document = openpyxl.load_workbook("Class.xlsx")

sheet = document.active
max_col = sheet.max_column
max_line = sheet.max_row
columTarget = 6
columLoop = 4
total_0_30 = 0
total_30_60 = 0
total_60_90 = 0
total_90_120 = 0
total_120_150 = 0
total = 0

print(max_line)
for i in range(4, 53):
    cell_obj = sheet.cell(row=i, column=columLoop)

    if cell_obj.value > 0 and cell_obj.value <= 30:
        total_0_30 = total_0_30 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 30 and cell_obj.value <= 60:
        total_30_60 = total_30_60 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 60 and cell_obj.value <= 90:
        total_60_90 = total_60_90 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 90 and cell_obj.value <= 120:
        total_90_120 = total_90_120 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 120 and cell_obj.value <= 150:
        total_120_150 = total_120_150 + sheet.cell(row=i, column=columTarget).value


print("0-100   = " , round(total_0_30*100),'%')
print("100-150 = " , round(total_30_60*100),'%')
print("150-200 = " , round(total_60_90*100),'%')
print("200-250 = " , round(total_90_120*100),'%')
print("250-300 = " , round(total_120_150*100),'%')
print()
print("Total % = ", round(
                total_0_30
                +total_30_60
                +total_60_90
                +total_90_120
                +total_120_150)*100)
