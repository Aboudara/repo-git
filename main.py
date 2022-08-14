# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl
import math

document = openpyxl.load_workbook("Class.xlsx")

sheet = document.active
max_col = sheet.max_column
max_line = sheet.max_row
total_0_500 = 0
total_500_1000 = 0
total_1000_2000 = 0
total_2000_3000 = 0
total_3000_4000 = 0
total_4000_5000 = 0
total = 0

for i in range(4, max_line+1):
    cell_obj = sheet.cell(row=i, column=5)
    if cell_obj.value > 0 and cell_obj.value <= 500:
        total_0_500 = total_0_500 + sheet.cell(row=i, column=6).value

    if cell_obj.value > 500 and cell_obj.value <= 1000:
        total_500_1000 = total_500_1000 + sheet.cell(row=i, column=6).value

    if cell_obj.value > 1000 and cell_obj.value <= 2000:
        total_1000_2000 = total_1000_2000 + sheet.cell(row=i, column=6).value

    if cell_obj.value > 2000 and cell_obj.value <= 3000:
        total_2000_3000 = total_2000_3000 + sheet.cell(row=i, column=6).value

    if cell_obj.value > 3000 and cell_obj.value <= 4000:
        total_3000_4000 = total_3000_4000 + sheet.cell(row=i, column=6).value

    if cell_obj.value > 4000 :
        total_4000_5000 = total_4000_5000 + sheet.cell(row=i, column=6).value


print("0-0.5 = " , math.ceil(total_0_500*100),'%')
print("0.5-1 = " , math.ceil(total_500_1000*100),'%')
print(" 1-2  = " , math.ceil(total_1000_2000*100),'%')
print(" 2-3  = " , math.ceil(total_2000_3000*100),'%')
print(" 3-4  = " , math.ceil(total_3000_4000*100),'%')
print(" > 4  = " , math.ceil(total_4000_5000*100),'%')

print("Total % = ", (total_0_500
                   +total_500_1000
                   +total_1000_2000
                   +total_2000_3000
                   +total_3000_4000
                   +total_4000_5000)*100)
