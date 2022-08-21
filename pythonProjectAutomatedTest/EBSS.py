# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl
from openpyxl import Workbook
import math

document = openpyxl.load_workbook("Analysis_EBSS_Test.xlsx")
#document = openpyxl.load_workbook("novembre.xlsx")

wb = Workbook()
print(wb.sheetnames)

sheet = document.active
max_col = sheet.max_column
max_line = sheet.max_row

print(max_col, max_line)
"""columTarget = 7
columLoop = 2

pourcent_0_100 = 0
pourcent_100_150 = 0
pourcent_150_200 = 0
pourcent_200_250 = 0
pourcent_250_300 = 0
pourcent_300_350 = 0
pourcent_350_400 = 0
totalPourcent = 0

NbParcel_0_100 = 0
NbParcel_100_150 = 0
NbParcel_150_200 = 0
NbParcel_200_250 = 0
NbParcel_250_300 = 0
NbParcel_300_350 = 0
NbParcel_350_400 = 0
totalNbParcel = 0

if columTarget == 6:
    for i in range(4, 53):
        cell_obj = sheet.cell(row=i, column=columLoop)

        if cell_obj.value > 0 and cell_obj.value <= 10:
            pourcent_0_100 = pourcent_0_100 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 10 and cell_obj.value <= 15:
            pourcent_100_150 = pourcent_100_150 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 15 and cell_obj.value <= 20:
            pourcent_150_200 = pourcent_150_200 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 20 and cell_obj.value <= 25:
            pourcent_200_250 = pourcent_200_250 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 25 and cell_obj.value <= 30:
            pourcent_250_300 = pourcent_250_300 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 30 and cell_obj.value <= 35:
            pourcent_300_350 = pourcent_300_350 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 35 and cell_obj.value <= 40:
            pourcent_350_400 = pourcent_350_400 + sheet.cell(row=i, column=columTarget).value


    print("0-100   = " , round(pourcent_0_100*100),'%')
    print("100-150 = " , round(pourcent_100_150*100),'%')
    print("150-200 = " , round(pourcent_150_200*100),'%')
    print("200-250 = " , round(pourcent_200_250*100),'%')
    print("250-300 = " , round(pourcent_250_300*100),'%')
    print("300-350 = " , round(pourcent_300_350*100),'%')
    print("350-400 = " , round(pourcent_350_400*100),'%')
    print()
    print("TotalPourcent % = ", round(
                    pourcent_0_100
                    +pourcent_100_150
                    +pourcent_150_200
                    +pourcent_200_250
                    +pourcent_250_300
                    +pourcent_300_350
                    +pourcent_350_400)*100)

if columTarget == 7:
        for i in range(4, 53):
            cell_obj = sheet.cell(row=i, column=columLoop)

            if cell_obj.value > 0 and cell_obj.value <= 10:
                NbParcel_0_100 = NbParcel_0_100 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 10 and cell_obj.value <= 15:
                NbParcel_100_150 = NbParcel_100_150 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 15 and cell_obj.value <= 20:
                NbParcel_150_200 = NbParcel_150_200 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 20 and cell_obj.value <= 25:
                NbParcel_200_250 = NbParcel_200_250 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 25 and cell_obj.value <= 30:
                NbParcel_250_300 = NbParcel_250_300 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 30 and cell_obj.value <= 35:
                NbParcel_300_350 = NbParcel_300_350 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 35 and cell_obj.value <= 40:
                NbParcel_350_400 = NbParcel_350_400 + sheet.cell(row=i, column=columTarget).value

        print("0-100   = ", NbParcel_0_100, 'Parcel')
        print("100-150 = ", NbParcel_100_150, 'Parcel')
        print("150-200 = ", NbParcel_150_200, 'Parcel')
        print("200-250 = ", NbParcel_200_250, 'Parcel')
        print("250-300 = ", NbParcel_250_300, 'Parcel')
        print("300-350 = ", NbParcel_300_350, 'Parcel')
        print("350-400 = ", NbParcel_350_400, 'Parcel')
        print()
        print("TotalNbParcel  = ", round(
            NbParcel_0_100
            + NbParcel_100_150
            + NbParcel_150_200
            + NbParcel_200_250
            + NbParcel_250_300
            + NbParcel_300_350
            + NbParcel_350_400))"""


