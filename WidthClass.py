# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl
import math

document = openpyxl.load_workbook("Class.xlsx")

sheet = document.active
max_col = sheet.max_column
max_line = sheet.max_row
columTarget = 6
columLoop = 3
total_0_100 = 0
total_100_150 = 0
total_150_200 = 0
total_200_250 = 0
total_250_300 = 0
total = 0

print(max_line)
for i in range(4, 53):
    cell_obj = sheet.cell(row=i, column=columLoop)

    if cell_obj.value > 0 and cell_obj.value <= 10:
        total_0_100 = total_0_100 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 10 and cell_obj.value <= 15:
        total_100_150 = total_100_150 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 15 and cell_obj.value <= 20:
        total_150_200 = total_150_200 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 20 and cell_obj.value <= 25:
        total_200_250 = total_200_250 + sheet.cell(row=i, column=columTarget).value

    if cell_obj.value > 25 and cell_obj.value <= 30:
        total_250_300 = total_250_300 + sheet.cell(row=i, column=columTarget).value


print("0-100   = " , round(total_0_100*100),'%')
print("100-150 = " , round(total_100_150*100),'%')
print("150-200 = " , round(total_150_200*100),'%')
print("200-250 = " , round(total_200_250*100),'%')
print("250-300 = " , round(total_250_300*100),'%')
print()
print("Total % = ", round(
                total_0_100
                +total_100_150
                +total_150_200
                +total_200_250
                +total_250_300)*100)
