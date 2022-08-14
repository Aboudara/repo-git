# FICHIERS EXCEL
# .XLSX
# openpyxl
import openpyxl
import math

document = openpyxl.load_workbook("Class.xlsx")
wb = openpyxl.Workbook()

sheet = document.active
max_col = sheet.max_column
max_line = sheet.max_row
columTarget = 7
columLoop = 5
pourcent_0_500 = 0
pourcent_500_1000 = 0
pourcent_1000_2000 = 0
pourcent_2000_3000 = 0
pourcent_3000_4000 = 0
totalpourcent = 0

NbParcel_0_500 = 0
NbParcel_500_1000 = 0
NbParcel_1000_2000 = 0
NbParcel_2000_3000 = 0
NbParcel_3000_4000 = 0
totalNbParcel = 0

if columTarget == 6:
    for i in range(4, 53):
        cell_obj = sheet.cell(row=i, column=columLoop)

        if cell_obj.value > 0 and cell_obj.value <= 500:
            pourcent_0_500 = pourcent_0_500 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 500 and cell_obj.value <= 1000:
            pourcent_500_1000 = pourcent_500_1000 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 1000 and cell_obj.value <= 2000:
            pourcent_1000_2000 = pourcent_1000_2000 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 2000 and cell_obj.value <= 3000:
            pourcent_2000_3000 = pourcent_2000_3000 + sheet.cell(row=i, column=columTarget).value

        if cell_obj.value > 3000 and cell_obj.value <= 4000:
            pourcent_3000_4000 = pourcent_3000_4000 + sheet.cell(row=i, column=columTarget).value



    print("0-0.5 = " , round(pourcent_0_500*100),'%')
    print("0.5-1 = " , round(pourcent_500_1000*100),'%')
    print(" 1-2  = " , round(pourcent_1000_2000*100),'%')
    print(" 2-3  = " , round(pourcent_2000_3000*100),'%')
    print(" 3-4  = " , round(pourcent_3000_4000*100),'%')

    print("pourcent % = ", round(pourcent_0_500
                       +pourcent_500_1000
                       +pourcent_1000_2000
                       +pourcent_2000_3000
                       +pourcent_3000_4000)*100)

if columTarget == 7:
        for i in range(4, 53):
            cell_obj = sheet.cell(row=i, column=columLoop)

            if cell_obj.value > 0 and cell_obj.value <= 500:
                NbParcel_0_500 = NbParcel_0_500 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 500 and cell_obj.value <= 1000:
                NbParcel_500_1000 = NbParcel_500_1000 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 1000 and cell_obj.value <= 2000:
                NbParcel_1000_2000 = NbParcel_1000_2000 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 2000 and cell_obj.value <= 3000:
                NbParcel_2000_3000 = NbParcel_2000_3000 + sheet.cell(row=i, column=columTarget).value

            if cell_obj.value > 3000 and cell_obj.value <= 4000:
                NbParcel_3000_4000 = NbParcel_3000_4000 + sheet.cell(row=i, column=columTarget).value

        print("0-0.5 = ", round(NbParcel_0_500), 'parcel')
        print("0.5-1 = ", round(NbParcel_500_1000), 'parcel')
        print(" 1-2  = ", round(NbParcel_1000_2000), 'parcel')
        print(" 2-3  = ", round(NbParcel_2000_3000), 'parcel')
        print(" 3-4  = ", round(NbParcel_3000_4000), 'parcel')
        print()
        print("NbParcel total= ", round(NbParcel_0_500
                                     + NbParcel_500_1000
                                     + NbParcel_1000_2000
                                     + NbParcel_2000_3000
                                     + NbParcel_3000_4000))


#--------------------------------------------------------------------------------------
# Get workbook active sheet
# from the active attribute
sheet = wb.active

# Cell objects also have row, column
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.
c1 = sheet.cell(row=1, column=1)
c2 = sheet.cell(row=1, column=2)
c3 = sheet.cell(row=1, column=3)

# writing values to cells
c1.value = "Class Weigth (kg)"
c2.value = "Rate"
c3.value = "Cumulative"

# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
la2 = sheet['A2']
la2.value = "0-0.5"
la3 = sheet['A3']
la3.value = "0.5-1"
la4 = sheet['A4']
la5 = sheet['A5']
la6 = sheet['A6']






la4.value = "1-2"


la5.value = "2-3"


la6.value = "3-4"

# B2 means column = 2 & row = 2.
lb2 = sheet['B2']
lb2.value = pourcent_0_500

lb3 = sheet['B3']
lb3.value = pourcent_500_1000

lb4 = sheet['B4']
lb4.value = pourcent_1000_2000

lb5 = sheet['B5']
lb5.value = pourcent_2000_3000

lb6 = sheet['B6']
lb6.value = pourcent_3000_4000

# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save("C:\\Users\\acgue\\Desktop\\AutomatedTest\\FileExcel\\excel\\demo.xlsx")